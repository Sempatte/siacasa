# admin_panel/training/training_controller.py
import os
import logging
from datetime import datetime, timedelta
import uuid
from io import BytesIO

from flask import Blueprint, render_template, request, redirect, url_for, flash, session, current_app, jsonify, send_from_directory, send_file
from werkzeug.utils import secure_filename

from admin_panel.auth.auth_middleware import login_required
from admin_panel.training.training_service import TrainingService
from bot_siacasa.infrastructure.db.neondb_connector import NeonDBConnector
from bot_siacasa.infrastructure.ai.training_manager import TrainingManager
from openai import OpenAI

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
except ImportError:  # pragma: no cover
    Workbook = None

logger = logging.getLogger(__name__)

# Crear blueprint para entrenamiento
training_blueprint = Blueprint('training', __name__)

# Inicializar servicios
db_connector = NeonDBConnector()
training_service = TrainingService(db_connector)
training_manager = TrainingManager(db_connector)

# Extensiones permitidas
ALLOWED_EXTENSIONS = {'txt', 'pdf', 'doc', 'docx', 'xls', 'xlsx', 'csv', 'json', 'md', 'html'}

def allowed_file(filename):
    """
    Verifica si un archivo tiene una extensión permitida.
    """
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def _format_datetime(dt: datetime) -> str:
    return dt.strftime('%d/%m/%Y %H:%M') if dt else '—'


def _format_duration(start: datetime, end: datetime) -> str:
    if not start or not end:
        return "En progreso"
    delta = end - start
    total_seconds = int(delta.total_seconds())
    if total_seconds < 60:
        return f"{total_seconds}s"
    minutes = total_seconds / 60
    if minutes < 60:
        return f"{minutes:.1f} min"
    hours = minutes / 60
    return f"{hours:.1f} h"

@training_blueprint.route('/')
@login_required
def index():
    """
    Muestra la página principal de gestión de entrenamiento.
    """
    bank_code = session.get('bank_code')
    bank_name = session.get('bank_name')
    
    # Obtener lista de archivos de entrenamiento
    training_files = training_service.get_training_files(bank_code)
    
    # Obtener historial de entrenamientos
    training_history = training_service.get_training_history(bank_code)
    today = datetime.now().date()
    default_start = (today - timedelta(days=7)).isoformat()
    default_end = today.isoformat()
    
    return render_template(
        'training.html',
        bank_name=bank_name,
        training_files=training_files,
        training_history=training_history,
        default_start_date=default_start,
        default_end_date=default_end,
        user_name=session.get('user_name', 'Admin')
    )

@training_blueprint.route('/upload', methods=['POST'])
@login_required
def upload_file():
    """
    Procesa la subida de un archivo de entrenamiento.
    """
    try:
        # Verificar si se envió un archivo
        if 'file' not in request.files:
            flash('No se seleccionó ningún archivo.', 'danger')
            return redirect(url_for('training.index'))
        
        file = request.files['file']
        
        # Verificar si se seleccionó un archivo
        if file.filename == '':
            flash('No se seleccionó ningún archivo.', 'danger')
            return redirect(url_for('training.index'))
        
        # Verificar si el archivo tiene una extensión permitida
        if not allowed_file(file.filename):
            flash(f'Tipo de archivo no permitido. Por favor, sube un archivo con alguna de estas extensiones: {", ".join(ALLOWED_EXTENSIONS)}', 'danger')
            return redirect(url_for('training.index'))
            
        # Obtener la descripción del archivo
        description = request.form.get('description', '')
        
        # Guardar archivo
        user_id = session.get('user_id')
        bank_code = session.get('bank_code')
        
        file_info = training_service.save_training_file(file, description, user_id, bank_code)
        
        flash(f'Archivo "{file.filename}" subido correctamente.', 'success')
        
        # Procesar el archivo inmediatamente
        try:
            success, message = training_manager.process_training_file(file_info['id'])
            if success:
                flash(f'Archivo procesado correctamente.', 'success')
            else:
                flash(f'El archivo se subió pero hubo un error al procesarlo: {message}', 'warning')
        except Exception as e:
            logger.error(f"Error al procesar archivo: {e}", exc_info=True)
            flash('El archivo se subió pero hubo un error al procesarlo.', 'warning')
        
        # Redireccionar a la página de entrenamiento
        return redirect(url_for('training.index'))
    
    except Exception as e:
        logger.error(f"Error al subir archivo: {e}", exc_info=True)
        flash('Error al subir el archivo. Por favor, intenta de nuevo.', 'danger')
        return redirect(url_for('training.index'))


@training_blueprint.route('/export', methods=['POST'])
@login_required
def export_training_report():
    """
    Genera un reporte en Excel con el estado de los entrenamientos.
    """
    if Workbook is None:
        flash('No se puede generar el reporte porque falta la librería openpyxl.', 'danger')
        return redirect(url_for('training.index'))

    try:
        start_date_str = request.form.get('start_date')
        end_date_str = request.form.get('end_date')

        today = datetime.now()
        default_start = today - timedelta(days=7)

        start_dt = datetime.strptime(start_date_str, '%Y-%m-%d') if start_date_str else default_start
        end_dt = datetime.strptime(end_date_str, '%Y-%m-%d') if end_date_str else today

        # Asegurar rango válido y cubrir todo el día final
        if start_dt > end_dt:
            flash('La fecha de inicio no puede ser mayor que la fecha fin.', 'danger')
            return redirect(url_for('training.index'))

        start_dt = start_dt.replace(hour=0, minute=0, second=0, microsecond=0)
        end_dt = end_dt.replace(hour=23, minute=59, second=59, microsecond=999999)

        bank_code = session.get('bank_code')
        bank_name = session.get('bank_name', 'Banco')

        sessions = training_service.get_training_sessions_by_range(bank_code, start_dt, end_dt)
        if not sessions:
            flash('No se encontraron sesiones de entrenamiento en el rango seleccionado.', 'warning')
            return redirect(url_for('training.index'))

        session_files = training_service.get_session_files_by_range(bank_code, start_dt, end_dt)

        workbook = Workbook()
        sheet_sessions = workbook.active
        sheet_sessions.title = "Sesiones"

        header = [
            "Fecha inicio",
            "Fecha fin",
            "Duración",
            "Estado",
            "Archivos totales",
            "Correctos",
            "Con error",
            "Iniciado por"
        ]
        sheet_sessions.append(header)

        bold_font = Font(bold=True)
        for cell in sheet_sessions[1]:
            cell.font = bold_font
            cell.alignment = Alignment(horizontal="center")

        for session_row in sessions:
            sheet_sessions.append([
                _format_datetime(session_row.get('start_time')),
                _format_datetime(session_row.get('end_time')),
                _format_duration(session_row.get('start_time'), session_row.get('end_time')),
                session_row.get('status', '').capitalize(),
                session_row.get('files_count', 0),
                session_row.get('success_count', 0),
                session_row.get('error_count', 0),
                session_row.get('initiated_by_name') or '—'
            ])

        # Ajustar ancho básico
        for column_cells in sheet_sessions.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            adjusted_width = max(15, max_length + 2)
            sheet_sessions.column_dimensions[column_cells[0].column_letter].width = adjusted_width

        sheet_files = workbook.create_sheet("Archivos por sesión")
        files_header = [
            "Sesión",
            "Fecha inicio",
            "Estado sesión",
            "Archivo",
            "Tipo",
            "Tamaño (KB)",
            "Estado archivo",
            "Mensaje de error",
            "Procesado el"
        ]
        sheet_files.append(files_header)
        for cell in sheet_files[1]:
            cell.font = bold_font
            cell.alignment = Alignment(horizontal="center")

        for detail in session_files:
            size_kb = round((detail.get('file_size') or 0) / 1024, 1)
            sheet_files.append([
                detail.get('session_id'),
                _format_datetime(detail.get('start_time')),
                detail.get('session_status', '').capitalize(),
                detail.get('original_filename'),
                detail.get('file_type'),
                size_kb,
                detail.get('file_status', '').capitalize(),
                detail.get('error_message') or '',
                _format_datetime(detail.get('processed_at'))
            ])

        for column_cells in sheet_files.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            adjusted_width = max(15, max_length + 2)
            sheet_files.column_dimensions[column_cells[0].column_letter].width = adjusted_width

        metadata_sheet = workbook.create_sheet("Resumen")
        metadata_sheet.append(["Banco", bank_name])
        metadata_sheet.append(["Código", bank_code])
        metadata_sheet.append(["Rango", f"{start_dt.date()} a {end_dt.date()}"])
        metadata_sheet.append(["Sesiones exportadas", len(sessions)])
        metadata_sheet.append(["Fecha de generación", _format_datetime(datetime.now())])
        for cell in metadata_sheet['A'][:5]:
            cell.font = bold_font

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = f"reporte_entrenamiento_{bank_code}_{start_dt.date()}_{end_dt.date()}.xlsx"
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            download_name=filename,
            as_attachment=True
        )

    except ValueError:
        flash('Formato de fecha inválido. Usa el formato AAAA-MM-DD.', 'danger')
        return redirect(url_for('training.index'))
    except Exception as e:
        logger.error(f"Error al exportar reporte de entrenamiento: {e}", exc_info=True)
        flash('Ocurrió un error al generar el reporte. Inténtalo más tarde.', 'danger')
        return redirect(url_for('training.index'))
@training_blueprint.route('/file/<file_id>/process', methods=['POST'])
@login_required
def process_file(file_id):
    """
    Procesa un archivo de entrenamiento manualmente.
    """
    try:
        # Obtener información del archivo
        file_info = training_service.get_file_details(file_id)
        
        if not file_info:
            flash('Archivo no encontrado.', 'danger')
            return redirect(url_for('training.index'))
        
        # Verificar que el archivo pertenece al banco del usuario
        if file_info['bank_code'] != session.get('bank_code'):
            flash('No tienes permisos para procesar este archivo.', 'danger')
            return redirect(url_for('training.index'))
        
        # Procesar archivo
        success, message = training_manager.process_training_file(file_id)
        
        if success:
            flash(f'Archivo procesado correctamente.', 'success')
        else:
            flash(f'Error al procesar archivo: {message}', 'danger')
        
        return redirect(url_for('training.file_details', file_id=file_id))
    
    except Exception as e:
        logger.error(f"Error al procesar archivo: {e}", exc_info=True)
        flash('Error al procesar el archivo. Por favor, intenta de nuevo.', 'danger')
        return redirect(url_for('training.index'))


@training_blueprint.route('/file/<file_id>')
@login_required
def file_details(file_id):
    """
    Muestra los detalles de un archivo de entrenamiento.
    """
    file_info = training_service.get_file_details(file_id)
    
    if not file_info:
        flash('Archivo no encontrado.', 'danger')
        return redirect(url_for('training.index'))
    
    # Verificar que el archivo pertenece al banco del usuario
    if file_info['bank_code'] != session.get('bank_code'):
        flash('No tienes permisos para ver este archivo.', 'danger')
        return redirect(url_for('training.index'))
    
    return render_template(
        'file_details.html',
        file=file_info,
        bank_name=session.get('bank_name'),
        user_name=session.get('user_name', 'Admin')
    )

@training_blueprint.route('/file/<file_id>/download')
@login_required
def download_file(file_id):
    """
    Descarga un archivo de entrenamiento.
    """
    file_info = training_service.get_file_details(file_id)
    
    if not file_info:
        flash('Archivo no encontrado.', 'danger')
        return redirect(url_for('training.index'))
    
    # Verificar que el archivo pertenece al banco del usuario
    if file_info['bank_code'] != session.get('bank_code'):
        flash('No tienes permisos para descargar este archivo.', 'danger')
        return redirect(url_for('training.index'))
    
    # Obtener el nombre del archivo en el sistema
    filename = file_info.get('filename')
    
    if not filename:
        flash('Error al descargar el archivo.', 'danger')
        return redirect(url_for('training.index'))
    
    # Obtener la carpeta de uploads
    upload_folder = current_app.config['UPLOAD_FOLDER']
    
    # Devolver el archivo
    return send_from_directory(
        upload_folder,
        filename,
        as_attachment=True,
        download_name=file_info['original_filename']
    )

@training_blueprint.route('/file/<file_id>/delete', methods=['POST'])
@login_required
def delete_file(file_id):
    """
    Elimina un archivo de entrenamiento.
    """
    file_info = training_service.get_file_details(file_id)
    
    if not file_info:
        flash('Archivo no encontrado.', 'danger')
        return redirect(url_for('training.index'))
    
    # Verificar que el archivo pertenece al banco del usuario
    if file_info['bank_code'] != session.get('bank_code'):
        flash('No tienes permisos para eliminar este archivo.', 'danger')
        return redirect(url_for('training.index'))
    
    # Eliminar archivo
    success = training_service.delete_file(file_id)
    
    if success:
        flash('Archivo eliminado correctamente.', 'success')
    else:
        flash('Error al eliminar el archivo.', 'danger')
    
    return redirect(url_for('training.index'))

@training_blueprint.route('/start', methods=['POST'])
@login_required
def start_training():
    """
    Inicia un entrenamiento con los archivos seleccionados.
    """
    # Obtener IDs de archivos seleccionados
    file_ids = request.form.getlist('file_ids[]')
    
    if not file_ids:
        flash('Debes seleccionar al menos un archivo para entrenar.', 'danger')
        return redirect(url_for('training.index'))
    
    try:
        # Iniciar entrenamiento
        user_id = session.get('user_id')
        bank_code = session.get('bank_code')
        
        session_id = training_service.start_training(file_ids, user_id, bank_code)
        
        flash('Entrenamiento iniciado correctamente.', 'success')
        
        # Redireccionar a la página de detalles del entrenamiento
        return redirect(url_for('training.session_details', session_id=session_id))
    
    except Exception as e:
        logger.error(f"Error al iniciar entrenamiento: {e}", exc_info=True)
        flash('Error al iniciar el entrenamiento. Por favor, intenta de nuevo.', 'danger')
        return redirect(url_for('training.index'))

@training_blueprint.route('/session/<session_id>')
@login_required
def session_details(session_id):
    """
    Muestra los detalles de una sesión de entrenamiento.
    """
    session_info = training_service.get_training_session_details(session_id)
    
    if not session_info:
        flash('Sesión de entrenamiento no encontrada.', 'danger')
        return redirect(url_for('training.index'))
    
    # Verificar que la sesión pertenece al banco del usuario
    if session_info['bank_code'] != session.get('bank_code'):
        flash('No tienes permisos para ver esta sesión.', 'danger')
        return redirect(url_for('training.index'))
    
    return render_template(
        'session_details.html',
        session=session_info,
        bank_name=session.get('bank_name'),
        user_name=session.get('user_name', 'Admin')
    )

@training_blueprint.route('/status')
@login_required
def training_status():
    """
    Devuelve el estado actual del entrenamiento (para actualización AJAX).
    """
    bank_code = session.get('bank_code')
    
    # Obtener estado de los entrenamientos activos
    active_sessions = db_connector.fetch_all(
        """
        SELECT id, status, files_count, success_count, error_count, start_time, end_time
        FROM training_sessions
        WHERE bank_code = %s AND status = 'processing'
        ORDER BY start_time DESC
        LIMIT 5
        """,
        (bank_code,)
    )
    
    return jsonify({
        'active_sessions': active_sessions,
        'timestamp': datetime.now().isoformat()
    })

@training_blueprint.route('/test_embedding', methods=['POST'])
@login_required
def test_embedding():
    """
    Prueba el funcionamiento de embeddings con una pregunta.
    """
    query = request.form.get('query', '')
    
    if not query:
        flash('Debes escribir una pregunta para probar.', 'danger')
        return redirect(url_for('training.index'))
    
    try:
        # Crear embedding para la consulta
        embedding = training_manager._get_embedding(query)
        
        if not embedding:
            flash('Error al generar embedding para la consulta.', 'danger')
            return redirect(url_for('training.index'))
        
        # Buscar documentos similares
        bank_code = session.get('bank_code')
        embedding_str = f"[{','.join(map(str, embedding))}]"
        
        similar_texts = db_connector.fetch_all(
            """
            SELECT 
                text, 
                file_id,
                1 - (embedding <=> %s::vector) as similarity
            FROM 
                text_embeddings
            WHERE 
                bank_code = %s
            ORDER BY 
                similarity DESC
            LIMIT 3
            """,
            (embedding_str, bank_code)
        )
        
        # Obtener información de los archivos relacionados
        if similar_texts:
            file_ids = [text['file_id'] for text in similar_texts]
            file_ids_str = ','.join([f"'{file_id}'" for file_id in file_ids])
            
            files_info = db_connector.fetch_all(
                f"""
                SELECT 
                    id, original_filename
                FROM 
                    training_files
                WHERE 
                    id IN ({file_ids_str})
                """
            )
            
            # Crear diccionario de archivos para acceso rápido
            files_dict = {file['id']: file for file in files_info}
            
            # Añadir nombre de archivo a los resultados
            for text in similar_texts:
                text['filename'] = files_dict.get(text['file_id'], {}).get('original_filename', 'Desconocido')
        
        # Generar respuesta usando OpenAI
        context = "\n\n".join([text['text'] for text in similar_texts])
        
        prompt = f"""
        Basándote en el siguiente contexto, responde a la pregunta del usuario.
        Si la información necesaria no está en el contexto, indica que no tienes suficiente información.

        CONTEXTO:
        {context}

        PREGUNTA:
        {query}
        """
        
        client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
        response = client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "Eres un asistente virtual bancario que responde basándose solo en el contexto proporcionado."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=400
        )
        
        answer = response.choices[0].message.content
        
        # Renderiazar resultados
        return render_template(
            'embedding_results.html',
            query=query,
            similar_texts=similar_texts,
            answer=answer,
            bank_name=session.get('bank_name'),
            user_name=session.get('user_name', 'Admin')
        )
    
    except Exception as e:
        logger.error(f"Error al probar embeddings: {e}", exc_info=True)
        flash('Error al probar embeddings. Por favor, intenta de nuevo.', 'danger')
        return redirect(url_for('training.index'))
